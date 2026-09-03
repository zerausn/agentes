"""
fill_docs_full.py — Vuelca ig_full_data.json al Excel y al .md completos.

Lee output/ig_full_data.json (generado por scrape_ig_full.py) y:
 1. Llena (o agrega) una fila por cada post en la hoja IG_469_Completo del Excel
 2. Actualiza/agrega cada post en REDHAC_Instagram.md

No borra filas existentes — solo actualiza o añade.
"""

import json
import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path
import re

BASE_DIR  = Path("/media/zerausn/D69493CF9493B08B/Users/ZN-/Documents/UNAD/"
                 "CURSOS/6/METODOLOGÍA Y GESTIÓN DE LA INVESTIGACIÓN/1/Documentacion/1")
EXCEL_PATH = BASE_DIR / "REDHAC_FINAL.xlsx"
IG_MD_PATH = BASE_DIR / "REDHAC_Instagram.md"
FULL_JSON  = Path(__file__).parent.parent / "output" / "ig_full_data.json"

# Columnas de la hoja IG_469_Completo (en orden)
COLS = [
    "#", "Link Post", "Texto Completo", "Total Likes", "Quien dio Like",
    "Nro Comentarios", "Comentarios (autor: texto)", "Compartidos",
    "Reposteo", "Fecha", "Menciones", "Imagen", "Link clickeable",
]


def load_full_data() -> dict:
    if not FULL_JSON.exists():
        print(f"Error: no existe {FULL_JSON}")
        print("Ejecuta primero: python3 scripts/scrape_ig_full.py")
        return {}
    return json.loads(FULL_JSON.read_text(encoding="utf-8"))


def update_excel(full_data: dict):
    print(f"\n=== Actualizando Excel: {EXCEL_PATH.name} ===")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["IG_469_Completo"]

    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    # Índice de filas existentes por código de post
    existing_rows: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2):
        link_cell = row[headers.get("Link Post", 2) - 1]
        if link_cell.value:
            code = str(link_cell.value).rstrip("/").split("/")[-1]
            existing_rows[code] = link_cell.row

    wrap = Alignment(wrap_text=True, vertical="top")

    def set_cell(ws, row_num, col_name, value):
        col_idx = headers.get(col_name)
        if col_idx is None:
            return
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = value
        cell.alignment = wrap

    updated = 0
    added = 0
    next_row = ws.max_row + 1
    num = ws.max_row - 1  # número secuencial base

    for code, data in full_data.items():
        if "error" in data and len(data) == 2:
            continue  # skip posts con error de scraping

        href = data.get("href", "")

        # Calcular rutas locales para imágenes/videos en base al código
        local_media = []
        for i in range(1, len(data.get("imgs", [])) + 1):
            local_media.append(f"media/REDHAC_{code}_foto{i}.jpg")
        for i in range(1, len(data.get("vids", [])) + 1):
            local_media.append(f"media/REDHAC_{code}_video{i}.mp4")
            
        imgs_str = "\n".join(local_media)
        likers_str = ", ".join(data.get("likers", []))
        menciones_str = ", ".join(data.get("menciones", []))
        comentarios_str = "\n".join(data.get("comentarios", []))
        fecha = data.get("fecha", "")

        if code in existing_rows:
            row_num = existing_rows[code]
            # Actualizar solo celdas vacías (no sobreescribir lo que ya tenía)
            row_vals = {
                ws.cell(row=row_num, column=c).value
                for c in range(1, ws.max_column + 1)
            }

            def update_if_empty(col_name, value):
                col_idx = headers.get(col_name)
                if col_idx is None:
                    return
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = value
                    cell.alignment = wrap

            update_if_empty("Total Likes", data.get("likes"))
            update_if_empty("Quien dio Like", likers_str)
            update_if_empty("Nro Comentarios", data.get("nro_comentarios"))
            update_if_empty("Comentarios (autor: texto)", comentarios_str)
            update_if_empty("Reposteo", data.get("reposteo", ""))
            update_if_empty("Fecha", fecha)
            update_if_empty("Menciones", menciones_str)
            update_if_empty("Imagen", imgs_str)
            update_if_empty("Texto Completo", data.get("texto", ""))
            updated += 1
        else:
            # Agregar nueva fila
            num += 1
            row_num = next_row
            next_row += 1

            set_cell(ws, row_num, "#", num)
            set_cell(ws, row_num, "Link Post", href)
            set_cell(ws, row_num, "Texto Completo", data.get("texto", ""))
            set_cell(ws, row_num, "Total Likes", data.get("likes"))
            set_cell(ws, row_num, "Quien dio Like", likers_str)
            set_cell(ws, row_num, "Nro Comentarios", data.get("nro_comentarios"))
            set_cell(ws, row_num, "Comentarios (autor: texto)", comentarios_str)
            set_cell(ws, row_num, "Compartidos", data.get("compartidos"))
            set_cell(ws, row_num, "Reposteo", data.get("reposteo", ""))
            set_cell(ws, row_num, "Fecha", fecha)
            set_cell(ws, row_num, "Menciones", menciones_str)
            set_cell(ws, row_num, "Imagen", imgs_str)
            set_cell(ws, row_num, "Link clickeable", href)
            added += 1

    wb.save(EXCEL_PATH)
    print(f"  ✓ {updated} filas actualizadas | {added} filas nuevas agregadas")
    print(f"  Total filas ahora: {ws.max_row - 1}")


def update_md(full_data: dict):
    print(f"\n=== Actualizando Markdown: {IG_MD_PATH.name} ===")
    if not IG_MD_PATH.exists():
        # Crear el .md desde cero
        lines = ["# REDHAC — Posts de Instagram\n\n"]
        for i, (code, data) in enumerate(full_data.items(), start=1):
            if "error" in data and len(data) == 2:
                continue
            href = data.get("href", "")
            
            local_media = []
            for j in range(1, len(data.get("imgs", [])) + 1):
                local_media.append(f"media/REDHAC_{code}_foto{j}.jpg")
            for j in range(1, len(data.get("vids", [])) + 1):
                local_media.append(f"media/REDHAC_{code}_video{j}.mp4")
                
            imgs_str = " | ".join(local_media) if local_media else ""
            menciones_str = ", ".join(data.get("menciones", []))
            comentarios_str = "\n".join(data.get("comentarios", []))
            lines.append(f"### {i}. [{href}]({href})\n")
            lines.append(f"**Likes:** {data.get('likes', '')} | **Fecha:** {data.get('fecha', '')}\n")
            if menciones_str:
                lines.append(f"**Menciones:** {menciones_str}\n")
            if imgs_str:
                lines.append(f"**Imagen(es):** {imgs_str}\n")
            lines.append(f"**Texto:** {data.get('texto', '')}\n")
            if comentarios_str:
                lines.append(f"**Comentarios:**\n{comentarios_str}\n")
            lines.append("\n---\n\n")
        IG_MD_PATH.write_text("".join(lines), encoding="utf-8")
        print(f"  ✓ .md creado con {len(full_data)} posts")
        return

    content = IG_MD_PATH.read_text(encoding="utf-8")
    updated = 0
    new_blocks = []

    for code, data in full_data.items():
        if "error" in data and len(data) == 2:
            continue
        href = data.get("href", "")
        
        local_media = []
        for j in range(1, len(data.get("imgs", [])) + 1):
            local_media.append(f"media/REDHAC_{code}_foto{j}.jpg")
        for j in range(1, len(data.get("vids", [])) + 1):
            local_media.append(f"media/REDHAC_{code}_video{j}.mp4")
            
        imgs_str = " | ".join(local_media) if local_media else ""
        menciones_str = ", ".join(data.get("menciones", []))

        escaped = re.escape(href.rstrip("/"))
        pattern = rf'(###\s+\d+\.\s+\[{escaped}[^\n]*\n(?:(?!###).)*?)(\n---)'

        img_line = f"**Imagen(es):** {imgs_str}\n" if imgs_str else ""
        mentions_line = f"**Menciones:** {menciones_str}\n" if menciones_str else ""

        def replace_block(m, img_line=img_line, mentions_line=mentions_line):
            block = m.group(1)
            sep = m.group(2)
            # Actualizar Imagen(es)
            if img_line:
                if "**Imagen" in block:
                    block = re.sub(r'\*\*Imagen[^*]*\*\*:[^\n]*\n?', img_line, block)
                else:
                    block = block.rstrip() + "\n" + img_line
            # Actualizar Menciones
            if mentions_line:
                if "**Menciones" in block:
                    block = re.sub(r'\*\*Menciones\*\*:[^\n]*\n?', mentions_line, block)
                else:
                    block = block.rstrip() + "\n" + mentions_line
            return block.rstrip() + "\n" + sep

        new_content, n = re.subn(pattern, replace_block, content, flags=re.DOTALL)
        if n > 0:
            content = new_content
            updated += 1
        else:
            # El post no está en el .md — agregar al final
            idx = content.count("### ")
            imgs_line = f"**Imagen(es):** {imgs_str}\n" if imgs_str else ""
            block = (
                f"\n### {idx+1}. [{href}]({href})\n"
                f"**Likes:** {data.get('likes', '')} | **Fecha:** {data.get('fecha', '')}\n"
                + (f"**Menciones:** {menciones_str}\n" if menciones_str else "")
                + imgs_line
                + f"**Texto:** {data.get('texto', '')}\n"
                + "\n---\n"
            )
            new_blocks.append(block)

    if new_blocks:
        content = content.rstrip() + "\n" + "".join(new_blocks)

    IG_MD_PATH.write_text(content, encoding="utf-8")
    print(f"  ✓ {updated} posts actualizados | {len(new_blocks)} posts nuevos agregados")


def main():
    full_data = load_full_data()
    if not full_data:
        return

    print(f"Posts disponibles en ig_full_data.json: {len(full_data)}")
    update_excel(full_data)
    update_md(full_data)
    print("\n✓ Documentos actualizados.")


if __name__ == "__main__":
    main()
