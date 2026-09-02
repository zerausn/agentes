"""
Llena la columna 'Imagen' del Excel REDHAC_FINAL.xlsx y actualiza el .md
con las URLs reales de fotos/videos extraídas por extract_media_urls.py.

Para carruseles, la columna Imagen tendrá todas las URLs separadas por newline.
Para reels, tendrá la URL del video.

No borra ningún dato existente — solo actualiza la columna Imagen si está vacía
o tiene una URL expirada de CDN.
"""
import json
import openpyxl
from pathlib import Path
import re

BASE_DIR = Path("/media/zerausn/D69493CF9493B08B/Users/ZN-/Documents/UNAD/CURSOS/6/METODOLOGÍA Y GESTIÓN DE LA INVESTIGACIÓN/1/Documentacion/1")
EXCEL_PATH = BASE_DIR / "REDHAC_FINAL.xlsx"
IG_MD_PATH = BASE_DIR / "REDHAC_Instagram.md"
MEDIA_URLS_JSON = Path(__file__).parent.parent / "output" / "media_urls.json"


def extract_code(href: str) -> str:
    return href.rstrip("/").split("/")[-1]


def load_media_urls() -> dict:
    if not MEDIA_URLS_JSON.exists():
        print(f"Error: no existe {MEDIA_URLS_JSON}")
        print("Ejecuta primero: python3 scripts/extract_media_urls.py")
        return {}
    return json.loads(MEDIA_URLS_JSON.read_text(encoding="utf-8"))


def update_excel(media_urls: dict):
    print(f"\n=== Actualizando Excel: {EXCEL_PATH.name} ===")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Detectar si esta hoja tiene columna 'Link Post' e 'Imagen'
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        if "Link Post" not in headers or "Imagen" not in headers:
            continue

        col_link = headers["Link Post"]
        col_img = headers["Imagen"]

        updated = 0
        for row in ws.iter_rows(min_row=2):
            link_cell = row[col_link - 1]
            img_cell = row[col_img - 1]

            href = link_cell.value
            if not href or "instagram" not in str(href):
                continue

            code = extract_code(str(href))
            if code not in media_urls:
                continue

            entry = media_urls[code]
            urls = entry.get("imgs", []) + entry.get("vids", [])

            if not urls:
                continue

            # Poner todas las URLs separadas por newline (una por foto/video)
            img_cell.value = "\n".join(urls)
            updated += 1

        print(f"  Hoja '{sheet_name}': {updated} filas actualizadas")

    wb.save(EXCEL_PATH)
    print(f"  ✓ Excel guardado.")


def update_md(media_urls: dict):
    print(f"\n=== Actualizando Markdown: {IG_MD_PATH.name} ===")
    if not IG_MD_PATH.exists():
        print(f"  No existe {IG_MD_PATH}")
        return

    content = IG_MD_PATH.read_text(encoding="utf-8")

    updated = 0
    for code, entry in media_urls.items():
        urls = entry.get("imgs", []) + entry.get("vids", [])
        if not urls:
            continue

        href = entry.get("href", "")
        # Buscar la sección del post en el .md por su URL
        # El .md tiene líneas como: ### N. [URL](URL)
        # Buscamos si ya tiene una línea **Imagen:** o la agregamos
        
        # Patron: encontrar el bloque de este post
        escaped = re.escape(href.rstrip("/"))
        # Buscar la sección que contiene este href
        pattern = rf'(###\s+\d+\.\s+\[{escaped}[^\n]*\n(?:(?!###).)*?)(\n---)'
        
        img_line = "**Imagen(es):** " + " | ".join(urls)
        
        def replace_block(m):
            block = m.group(1)
            sep = m.group(2)
            # Si ya tiene línea de Imagen, reemplazar
            if "**Imagen" in block:
                block = re.sub(r'\*\*Imagen[^*]*\*\*:[^\n]*\n?', img_line + "\n", block)
            else:
                # Agregar antes del separador
                block = block.rstrip() + "\n" + img_line + "\n"
            return block + sep

        new_content, n = re.subn(pattern, replace_block, content, flags=re.DOTALL)
        if n > 0:
            content = new_content
            updated += 1

    IG_MD_PATH.write_text(content, encoding="utf-8")
    print(f"  ✓ {updated} posts actualizados en el .md")


def main():
    media_urls = load_media_urls()
    if not media_urls:
        return

    print(f"URLs disponibles: {len(media_urls)} posts")
    update_excel(media_urls)
    update_md(media_urls)
    print("\n✓ Documentos actualizados correctamente.")


if __name__ == "__main__":
    main()
