"""
generate_excel.py - Genera reporte Excel REDHAC con 4 hojas
============================================================
Combina los datos extraídos de Facebook e Instagram para producir un
Excel de caracterización y propuesta de investigación ECACEN/UNAD.

Hojas generadas:
  1. Caracterización REDHAC → comparativa FB vs IG (seguidores, posts, método, etc.)
  2. FB_Posts_Extraidos     → posts de Facebook limpios con fecha estimada
  3. IG_469_Media           → 469 hrefs de Instagram con alt/caption/tipo
  4. Propuesta_ECACEN       → opciones A/B/C de investigación + detalle Opción A

Requiere archivos previos en output/:
  - ig_all_final.json    (generado por continue_ig.py)
  - fb_chrome_all.json   (generado por fb_chrome_full.py)

Salida:
  - output/REDHAC_Caracterizacion_y_Propuesta_ECACEN_2026.xlsx

Requisitos:
  pip install openpyxl

Uso:
  python3 scripts/generate_excel.py
"""

import json
import openpyxl
import re
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ─────────────────────────────────────────────────────────────────────
OUT_DIR = Path(__file__).parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)

ig_path = OUT_DIR / "ig_all_final.json"
fb_path = OUT_DIR / "fb_chrome_all.json"

ig_data = json.loads(ig_path.read_text(encoding="utf-8")) if ig_path.exists() else {"header": {}, "media": []}
fb_data = json.loads(fb_path.read_text(encoding="utf-8")) if fb_path.exists() else {"header": {}, "posts": []}

ig_header = ig_data.get("header", {})
ig_media = ig_data.get("media", [])
fb_header = fb_data.get("header", {})
fb_posts = fb_data.get("posts", [])

# ── Limpiar posts de Facebook ─────────────────────────────────────────────────
# Quitar cabecera garbled (letras sueltas con espacio) y posts muy cortos
clean_posts: list[str] = []
for post in fb_posts:
    if len(post) < 60:
        continue
    if "WhatsApp Seguir" in post and len(post) < 800:
        continue
    # Buscar inicio real del post (palabras clave del contenido de REDHAC)
    m = re.search(
        r'(Huerta Semillas|Nos vemos este domingo|LA ACACIA|Huerta comunitaria'
        r'|Nos vemos este domingo, en la Huerta)',
        post,
    )
    if m:
        clean = post[m.start():]
    else:
        # Si empieza con letras sueltas "d o s n e r t...", buscar primera palabra real
        if re.match(r'^([a-z] [a-z] )', post):
            parts = post.split(" ")
            idx = 0
            for i, word in enumerate(parts):
                if len(word) > 3 and word not in ["Seguir", "WhatsApp"]:
                    idx = i
                    break
            clean = " ".join(parts[idx:]) if idx else post
        else:
            clean = post

    clean = re.sub(r'\s+', ' ', clean).strip()
    if len(clean) > 40:
        clean_posts.append(clean)

# Fallback si el filtro eliminó demasiado
if len(clean_posts) < 3:
    clean_posts = [p for p in fb_posts if "Huerta" in p or "Frijolada" in p or "ACACIA" in p]

# ── Estilos ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
accent = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Calibri", bold=True, color="1F4E78", size=14)
subtitle_font = Font(name="Calibri", bold=True, color="1F4E78", size=11)
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)


def style_header_row(ws, row: int, cols: int):
    """Aplica estilo de encabezado azul oscuro a una fila completa."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


# ── Hoja 1: Caracterización REDHAC ───────────────────────────────────────────
ws1 = wb.active
ws1.title = "Caracterización REDHAC"
ws1.sheet_properties.pageSetUpPr.fitToPage = True
ws1.merge_cells('A1:D1')
ws1['A1'] = (
    "CARACTERIZACIÓN RED DE HUERTOS AGROECOLÓGICOS DE CALI (REDHAC) "
    "- Extracción vía Chrome con perfil Edge 02/09/2026"
)
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 22

ws1['A3'] = "Criterio"
ws1['B3'] = "Facebook"
ws1['C3'] = "Instagram"
ws1['D3'] = "Observaciones / Fuente"
style_header_row(ws1, 3, 4)
ws1.row_dimensions[3].height = 18

char_data = [
    ["URL", "facebook.com/Reddehuertosagroecologicosdecali", "instagram.com/redhuertosagroecali", "Bing redirect decodificado + CDP Edge"],
    ["Nombre", "Red De Huertos Agroecologicos Cali", "RedHuertosAgroecologicosCali", "og:title / header h2"],
    ["Seguidores", "1549 seguidores", "1325 seguidores", "FB: 1.549 Me gusta (og:description) | IG: header"],
    ["Seguidos", "82 seguidos", "217 seguidos", ""],
    ["Publicaciones", "6 visibles en DOM [role=article] (estimado 80-120 histórico)", "469 publicaciones (100% extraído)", "IG verificado contra header 469"],
    ["Categoría", "Agricultura", "Colectivo ambiental / ESAL", "FB: Página · Agricultura"],
    ["Contacto", "redhuertosagroecologicoscali@gmail.com", "redhuertosagroecali (DM)", "FB body innerText"],
    ["Dirección", "Calle13, Santiago de Cali, Colombia, 760032", "Cali (ubicación posts)", "FB Detalles"],
    ["ID interno", "fb://profile/100070371399388", "IG business 1325", "og:al:android:url"],
    ["Highlights", "Destacados: PazconlaNaturaleza", "Memoria | Mingas Huerterxs | Pedagogía | Aud Publica | Veeduria Ciudad | Conversatorios", "IG header"],
    ["Contenido predominante", "Mingas, ollas rodantes, frijoladas, huertas comunitarias, querellas (Villas de Guadalupe)", "Reels y fotos de mingas, pedagogía, veeduría, memoria", "Body innerText + alt"],
    ["Frecuencia", "Último 30 ene 2026 (Encuentro 1 feb El Morro)", "Alta, último 23 ago 2026", ""],
    ["Engagement", "+20 fotos por post, 1 persona asistió", "Alt con likes/comments", ""],
    ["Alianzas visibles", "DAGMA, CVC, UNAL Palmira, Univalle, SENA, ProPacífico", "Mismos + @ccinec__, @redagriculturasparalavida, @kelorengifo", ""],
    ["Método extracción", "Chrome 152 + Edge profile + CDP ws://127.0.0.1:9222 + scroll inner div (hasFeed) + body innerText clean U+034F", "Chrome + CDP scroll window 60 iteraciones, a[href*=\"/p/\"]", "API Meta bloqueada #10 sin Page Public Content Access"],
    ["Estado", "Parcial: 6 visibles (anti-scraping virtualizado)", "Completo: 469/469", ""],
]

for i, row in enumerate(char_data, start=4):
    cell1 = ws1.cell(row=i, column=1, value=row[0])
    cell1.font = Font(bold=True)
    cell1.alignment = Alignment(vertical='center', wrap_text=True)
    for j, val in enumerate(row[1:], start=2):
        c = ws1.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = thin_border
    ws1.row_dimensions[i].height = 28 if len(row[1]) > 60 else 18
    if i % 2 == 0:
        for col in range(1, 5):
            ws1.cell(row=i, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for i, width in enumerate([28, 42, 42, 38], start=1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# ── Hoja 2: FB Posts Extraídos ───────────────────────────────────────────────
ws2 = wb.create_sheet("FB_Posts_Extraidos")
ws2.merge_cells('A1:F1')
ws2['A1'] = "FACEBOOK - Posts extraídos (Chrome Edge profile) - 6 visibles + significativos"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

headers2 = ["#", "Fecha estimada", "Texto (limpio sin ofuscación)", "Links", "Likes/Comments", "Imagen"]
for col, h in enumerate(headers2, start=1):
    ws2.cell(row=2, column=col).value = h
style_header_row(ws2, 2, 6)
ws2.row_dimensions[2].height = 18

for idx, post in enumerate(clean_posts[:20], start=1):
    # Estimación de fecha por palabras clave en el contenido
    date = (
        "30 ene 2026" if "El Morro" in post
        else "12 feb 2023" if "ACACIA" in post
        else "01 feb 2026" if "Frijolada" in post
        else "2026-08"
    )
    if "Semillas del Futuro" in post:
        date = "2023-02-12 aprox"

    ws2.append([idx, date, post[:3500], "photo/?fbid=... / posts", "", "scontent.fclo8-1.fna.fbcdn.net"])
    row_num = ws2.max_row
    for col in range(1, 7):
        ws2.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        ws2.cell(row=row_num, column=col).border = thin_border
    ws2.row_dimensions[row_num].height = 60

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 85
ws2.column_dimensions['D'].width = 28
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 22
ws2.sheet_properties.pageSetUpPr.fitToPage = True

# ── Hoja 3: Instagram 469 Media ──────────────────────────────────────────────
ws3 = wb.create_sheet("IG_469_Media")
ws3.merge_cells('A1:E1')
ws3['A1'] = "INSTAGRAM @redhuertosagroecali - 469 publicaciones extraídas 100% (Chrome CDP)"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

headers3 = ["#", "Href", "Alt / Caption (primeros 400 chars)", "Imagen", "Tipo"]
for col, h in enumerate(headers3, start=1):
    ws3.cell(row=2, column=col).value = h
style_header_row(ws3, 2, 5)

for idx, media in enumerate(ig_media, start=1):
    alt = media.get("alt", "")[:500]
    href = media.get("href", "")
    img = media.get("img", "")[:120]
    typ = "reel" if "/reel/" in href else "post" if "/p/" in href else "other"
    ws3.append([idx, href, alt, img, typ])
    row_num = ws3.max_row
    for col in range(1, 6):
        ws3.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(row=row_num, column=col).border = thin_border
    ws3.row_dimensions[row_num].height = 40
    if idx >= 500:
        break

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 85
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 10
ws3.freeze_panes = 'A3'
ws3.auto_filter.ref = f"A2:E{ws3.max_row}"
ws3.sheet_properties.pageSetUpPr.fitToPage = True

# ── Hoja 4: Propuesta Investigación ECACEN ───────────────────────────────────
ws4 = wb.create_sheet("Propuesta_ECACEN")
ws4.merge_cells('A1:D1')
ws4['A1'] = "PROPUESTA DE INVESTIGACIÓN ECACEN - UNAD - Resultado de Aprendizaje 1"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')
ws4.merge_cells('A2:D2')
ws4['A2'] = "Posicionamiento: Interesadas / Vinculadas / Posicionadas - Programa RESALTAR CCC (Cámara de Comercio de Cali)"
ws4['A2'].font = Font(italic=True, color="595959")
ws4['A2'].alignment = Alignment(horizontal='center')

ws4['A4'] = "Programa RESALTAR: Más Recursos, Más Impacto (CCC + Makaia)"
ws4['A4'].font = subtitle_font
ws4['A4'].fill = light_blue
ws4.merge_cells('A5:D5')
ws4['A5'] = (
    "Para ESAL constituidas y activas. Diagnóstico de madurez en movilización de recursos → "
    "3 rutas: Interesadas (primeros pasos), Vinculadas (ya gestionan recursos, buscan consolidar), "
    "Posicionadas (ampliar alcance estratégico). Formación virtual + asesoría + conexión financiación. "
    "Requisitos: ESAL, operación activa, disponibilidad virtual. Contacto: resaltar@ccc.org.co"
)
ws4['A5'].alignment = Alignment(wrap_text=True, vertical='top')
ws4.row_dimensions[5].height = 38

for col, h in enumerate(["Opción", "Problema (para Anexo Único)", "Enfoque ECACEN", "Por qué posiciona en RESALTAR"], start=1):
    ws4.cell(row=7, column=col).value = h
    ws4.cell(row=7, column=col).fill = blue_fill
    ws4.cell(row=7, column=col).font = header_font
    ws4.cell(row=7, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4.cell(row=7, column=col).border = thin_border
ws4.row_dimensions[7].height = 30

opciones = [
    ["A - RECOMENDADA\nMovilización de recursos",
     "Débil diversificación y sostenibilidad en ESAL ambientales comunitarias tipo REDHAC, pese a oferta institucional como RESALTAR (CCC), DAGMA (362 huertas) y CVC, que exige diagnóstico de madurez y estrategia clara para pasar de recursos limitados a movilización sostenible.",
     "Gestión de las Organizaciones + Desarrollo Sostenible",
     "Apunta directo al objetivo de RESALTAR: pasar de dependencia limitada a estrategia diversificada. Permite diagnosticar a REDHAC como Interesada → Vinculada."],
    ["B - Formalización",
     "Bajo nivel de formalización y madurez organizacional de huertas urbanas comunitarias que limita el acceso a rutas de fortalecimiento como RESALTAR (niveles Interesadas/Vinculadas/Posicionadas), aunque cumplen criterios ESAL y operación activa en Cali.",
     "Emprendimiento Social y Solidario + Gestión",
     "Útil si REDHAC aún no es ESAL formal. Posiciona como Interesadas que necesitan constitución y madurez para acceder."],
    ["C - Articulación",
     "Desarticulación entre oferta de fortalecimiento empresarial (CCC Resaltar, DAGMA 362 huertas, CVC) y demanda de colectivos huerteros, que impide escalar bioinsumos (bocashi, compost) y ecoturismo pedagógico (Sendero Calima 4h/8h) ya validados.",
     "Cadenas Productivas Agroindustriales + Gestión",
     "Ideal para enfoque de cadena productiva y escalamiento. Posiciona como Vinculadas que ya gestionan recursos (bioinsumos) y buscan consolidar."],
]

for i, row in enumerate(opciones, start=8):
    ws4.cell(row=i, column=1, value=row[0]).font = Font(bold=True, color="1F4E78")
    ws4.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for j, val in enumerate(row[1:], start=2):
        c = ws4.cell(row=i, column=j, value=val)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = thin_border
    ws4.row_dimensions[i].height = 65
    if i == 8:  # Destacar opción A recomendada
        for col in range(1, 5):
            ws4.cell(row=i, column=col).fill = accent

# Detalle Propuesta A
ws4.merge_cells('A12:D12')
ws4['A12'] = "DETALLE PROPUESTA A (para diligenciar Anexo Único)"
ws4['A12'].font = subtitle_font
ws4['A12'].fill = light_blue

details = [
    ["Título", "Modelo de movilización y diversificación de recursos para la sostenibilidad de la Red de Huertos Agroecológicos de Cali (REDHAC) como ESAL ambiental comunitaria - Santiago de Cali, 2024-2025"],
    ["Línea ECACEN", "Gestión de las Organizaciones (principal) + Desarrollo Regional Sostenible + Emprendimiento Social y Solidario"],
    ["Pregunta", "¿Cómo incide la implementación de una estrategia de movilización y diversificación de recursos (diagnóstico RESALTAR) en la sostenibilidad organizacional de la REDHAC en Cali 2024-2025?"],
    ["Población", "REDHAC: 50 procesos caracterizados Univalle 2023 / 362 huertas DAGMA. Muestra: 30 líderes (Mojica, Villas de Guadalupe, El Morro Camilo Osuna) - accesible vía IG 1325 seguidores"],
    ["Variables", "VI: Nivel de madurez en movilización de recursos (Interesada/Vinculada/Posicionada) | VD: Sostenibilidad organizacional (diversificación, acceso a financiación)"],
    ["Objetivo General", "Proponer una estrategia de movilización y diversificación de recursos para la REDHAC que fortalezca su sostenibilidad y acceso a oportunidades tipo RESALTAR en Cali 2024-2025."],
    ["Objetivos Específicos", "1) Caracterizar nivel de madurez y prácticas actuales de movilización (encuesta RESALTAR + Univalle). 2) Identificar factores que limitan diversificación. 3) Analizar articulación con oferta institucional (CCC, DAGMA, CVC). 4) Diseñar ruta Interesadas→Vinculadas con manual de movilización."],
    ["Justificación", "Teórica: Aporta a Gestión y Desarrollo Sostenible. Práctica: 7,3M colombianos inseguridad alimentaria (FAO 2022), 6.500 familias CVC. Metodológica: Cualitativo + diagnóstico RESALTAR validado. Social: Permite a REDHAC pasar a Vinculadas y acceder a financiación."],
    ["Delimitación", "Espacial: Cali, 22 comunas + corregimientos. Temporal: 2024-2025 + campo 2026-1. Temática: Ciencias administrativas (no agronómica). Poblacional: Solo REDHAC."],
    ["Viabilidad", "Alta: Acceso vía IG/FB, sin impedimentos éticos (ESAL abierta, consentimiento), tiempo 1 semestre, fuentes secundarias abundantes (Univalle 50, DAGMA, CCC)."],
]

start_row = 13
for detail in details:
    ws4.cell(row=start_row, column=1, value=detail[0]).font = Font(bold=True)
    ws4.cell(row=start_row, column=1).alignment = Alignment(vertical='top')
    ws4.cell(row=start_row, column=1).border = thin_border
    ws4.merge_cells(f'B{start_row}:D{start_row}')
    c = ws4.cell(row=start_row, column=2, value=detail[1])
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = thin_border
    ws4.row_dimensions[start_row].height = 28 if len(detail[1]) < 150 else 38
    start_row += 1

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 38
ws4.sheet_properties.pageSetUpPr.fitToPage = True

# ── Configuración de impresión para todas las hojas ──────────────────────────
for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# ── Guardar ──────────────────────────────────────────────────────────────────
out_file = OUT_DIR / "REDHAC_Caracterizacion_y_Propuesta_ECACEN_2026.xlsx"
wb.save(out_file)
print(f"Excel guardado: {out_file}")
print(f"IG: {len(ig_media)} media | FB limpio: {len(clean_posts)} posts")
